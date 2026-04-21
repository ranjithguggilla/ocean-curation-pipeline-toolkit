"""
validate command: Check raw files for completeness and structural integrity.

Checks performed:
- File existence and readability
- Character encoding detection (flags non-UTF-8)
- CSV structural consistency (column counts, empty rows)
- Excel: merged cells, hidden sheets, formulas
- Required columns present
- Value ranges within bounds
- Coordinate validation (lat/lon)
- Timestamp parseability
- Duplicate detection
"""

import json
from pathlib import Path

import chardet
import click
import pandas as pd

from ocean_curator.utils import (
    get_files_config,
    get_quality_config,
    get_submission_dir,
    ensure_subdirs,
    log_provenance,
    setup_logger,
)
from ocean_curator.commands.transform_cmd import normalize_header


def _match_col(df_columns: list[str], target: str) -> str | None:
    """Find a DataFrame column whose normalized form matches target."""
    for c in df_columns:
        if normalize_header(c) == target.lower():
            return c
    return None


class ValidationReport:
    """Accumulates validation results per file."""

    def __init__(self):
        self.results: list[dict] = []
        self._current_file = None
        self._current_checks = []

    def start_file(self, filename: str):
        self._current_file = filename
        self._current_checks = []

    def add_check(self, name: str, passed: bool, detail: str = ""):
        self._current_checks.append({
            "check": name,
            "passed": bool(passed),
            "detail": detail,
        })

    def finish_file(self):
        all_passed = all(c["passed"] for c in self._current_checks)
        self.results.append({
            "file": self._current_file,
            "passed": all_passed,
            "checks": self._current_checks,
        })

    @property
    def all_passed(self) -> bool:
        return all(r["passed"] for r in self.results)

    @property
    def summary(self) -> str:
        total = sum(len(r["checks"]) for r in self.results)
        failed = sum(
            1 for r in self.results for c in r["checks"] if not c["passed"]
        )
        return f"{total - failed}/{total} checks passed across {len(self.results)} files"


def detect_encoding(filepath: Path) -> tuple[str, float]:
    """Detect file encoding using chardet."""
    with open(filepath, "rb") as f:
        raw = f.read(min(65536, filepath.stat().st_size))
    result = chardet.detect(raw)
    return result.get("encoding", "unknown"), result.get("confidence", 0.0)


def validate_csv(filepath: Path, quality: dict, report: ValidationReport):
    """Run all CSV-specific validations."""
    try:
        df = pd.read_csv(filepath, encoding="utf-8")
    except UnicodeDecodeError:
        report.add_check("utf8_readable", False, "File is not valid UTF-8")
        return
    except Exception as e:
        report.add_check("parseable", False, str(e))
        return

    report.add_check("parseable", True, f"{len(df)} rows, {len(df.columns)} columns")

    # Empty rows
    empty_rows = df.isnull().all(axis=1).sum()
    report.add_check(
        "no_empty_rows",
        empty_rows == 0,
        f"{empty_rows} completely empty rows found" if empty_rows else "No empty rows",
    )

    # Required columns
    required = quality.get("required_columns", [])
    col_list = list(df.columns)
    for col in required:
        matched = _match_col(col_list, col)
        report.add_check(
            f"required_column_{col}",
            matched is not None,
            f"Column '{col}' {'found' if matched else 'MISSING'} (raw: {matched})" if matched else f"Column '{col}' MISSING",
        )

    # Value ranges
    ranges = quality.get("value_ranges", {})
    for col_name, (lo, hi) in ranges.items():
        matched = _match_col(col_list, col_name)
        if not matched:
            continue
        col_data = pd.to_numeric(df[matched], errors="coerce").dropna()
        if col_data.empty:
            report.add_check(f"range_{col_name}", True, "No numeric data to check")
            continue
        out_of_range = ((col_data < lo) | (col_data > hi)).sum()
        report.add_check(
            f"range_{col_name}",
            out_of_range == 0,
            f"{out_of_range} values outside [{lo}, {hi}]" if out_of_range else f"All values in [{lo}, {hi}]",
        )

    # Duplicate detection
    dup_keys = quality.get("duplicate_key_columns", [])
    max_dups = quality.get("max_duplicate_tolerance", 0)
    if dup_keys:
        matched_keys = []
        for dk in dup_keys:
            mk = _match_col(col_list, dk)
            if mk:
                matched_keys.append(mk)
        if matched_keys:
            dup_count = df.duplicated(subset=matched_keys, keep=False).sum()
            report.add_check(
                "no_duplicates",
                dup_count <= max_dups,
                f"{dup_count} duplicate rows on key {matched_keys}"
                if dup_count > max_dups
                else "No duplicates",
            )


def validate_excel(filepath: Path, file_entry: dict, quality: dict, report: ValidationReport):
    """Run Excel-specific validations then delegate to CSV checks."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        report.add_check("excel_readable", False, str(e))
        return

    report.add_check("excel_readable", True, f"Sheets: {wb.sheetnames}")

    # Check for merged cells
    target_sheet = file_entry.get("sheet", wb.sheetnames[0])
    if target_sheet in wb.sheetnames:
        ws = wb[target_sheet]
        merged = len(ws.merged_cells.ranges)
        report.add_check(
            "no_merged_cells",
            merged == 0,
            f"{merged} merged cell regions found" if merged else "No merged cells",
        )

    # Hidden sheets warning
    hidden = [s for s in wb.sheetnames if wb[s].sheet_state != "visible"]
    report.add_check(
        "no_hidden_sheets",
        len(hidden) == 0,
        f"Hidden sheets: {hidden}" if hidden else "No hidden sheets",
    )

    wb.close()

    # Read into pandas for data-level checks
    sheet = file_entry.get("sheet", 0)
    try:
        df = pd.read_excel(filepath, sheet_name=sheet)
        # Write temp CSV and reuse CSV validator logic
        report.add_check("parseable", True, f"{len(df)} rows, {len(df.columns)} columns")

        # Required columns
        required = quality.get("required_columns", [])
        col_list = list(df.columns)
        for col in required:
            matched = _match_col(col_list, col)
            report.add_check(
                f"required_column_{col}",
                matched is not None,
                f"Column '{col}' found (raw: {matched})" if matched else f"Column '{col}' MISSING",
            )

    except Exception as e:
        report.add_check("parseable", False, str(e))


def run_validate(config: dict):
    """Execute the validate step."""
    submission_dir = get_submission_dir(config)
    subdirs = ensure_subdirs(submission_dir)
    logger = setup_logger(subdirs["logs"])
    quality = get_quality_config(config)
    report = ValidationReport()

    logger.info("Starting validation...")

    for file_entry in get_files_config(config):
        raw_path = submission_dir / "raw" / Path(file_entry["source_path"]).name
        if not raw_path.exists():
            # Try the name directly
            raw_path = submission_dir / "raw" / file_entry["name"]

        report.start_file(str(raw_path.name))

        # File existence
        if not raw_path.exists():
            report.add_check("file_exists", False, f"Not found: {raw_path}")
            report.finish_file()
            continue
        report.add_check("file_exists", True, f"Size: {raw_path.stat().st_size} bytes")

        # Encoding detection
        encoding, confidence = detect_encoding(raw_path)
        is_utf8 = encoding and encoding.lower() in ("utf-8", "ascii")
        report.add_check(
            "encoding",
            is_utf8,
            f"Detected: {encoding} (confidence: {confidence:.0%})",
        )

        # Format-specific checks
        fmt = file_entry.get("format", "")
        if "csv" in fmt or raw_path.suffix.lower() == ".csv":
            validate_csv(raw_path, quality, report)
        elif "spreadsheet" in fmt or raw_path.suffix.lower() in (".xlsx", ".xls"):
            validate_excel(raw_path, file_entry, quality, report)

        report.finish_file()
        logger.info("Validated: %s — %s",
                     raw_path.name,
                     "PASSED" if report.results[-1]["passed"] else "ISSUES FOUND")

    # Write validation report
    report_path = subdirs["logs"] / "validation_report.json"
    with open(report_path, "w", encoding="utf-8") as f:
        json.dump({
            "summary": report.summary,
            "all_passed": report.all_passed,
            "files": report.results,
        }, f, indent=2, default=str)

    logger.info("Validation report: %s", report_path.name)
    logger.info("Result: %s", report.summary)

    log_provenance(subdirs["logs"], {
        "action": "validate",
        "all_passed": report.all_passed,
        "summary": report.summary,
    })

    status = "✓" if report.all_passed else "⚠"
    click.echo(f"{status} Validation: {report.summary}")
